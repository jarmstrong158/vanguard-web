"""
update_labels.py
----------------
Refreshes the label maker's SKU drop-down and descriptions from your newest
local master barcoding workbook. Run this whenever new codes are added to the
master (the file that gets emailed around and saved on each PC).

Workflow:
    1. Save the latest master barcoding file (e.g. in Downloads).
    2. Run this  ->  updates Barcoding_Label_Maker.xlsx (drop-down + descriptions).
    3. Open Barcoding_Label_Maker.xlsx, pick SKUs + bins, save.
    4. Run make_labels.py to print.

You only need to run this when the master changes -- not every time you print.

Usage:
    Double-click this file, or run:
        python update_labels.py
        python update_labels.py --master "C:\\path\\to\\Barcoding_021526.xlsx"

The first run auto-installs the one package it needs (openpyxl).
"""
import argparse
import importlib
import re
import subprocess
import sys
from pathlib import Path


def ensure_packages():
    """Make sure openpyxl is importable; install it if not."""
    try:
        importlib.import_module("openpyxl")
        return
    except ImportError:
        pass
    print("First-time setup: installing openpyxl (needs internet, once)...")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install",
                               "--quiet", "openpyxl"])
    except Exception as exc:
        print(f"\nCould not install openpyxl automatically. Reason: {exc}")
        print("Fix: open a Command Prompt and run, then try again:")
        print(f'  "{sys.executable}" -m pip install openpyxl')
        input("\nPress Enter to close...")
        sys.exit(1)
    importlib.invalidate_caches()


ensure_packages()

import openpyxl

WORKBOOK_NAME = "Barcoding_Label_Maker.xlsx"
MASTER_HINT = "barcod"                 # master files have this in their filename
CONFIG_NAME = "master_location.txt"    # remembers where the master was found


def looks_like_master(path):
    """True if the workbook has the master's data tabs (List / Reference)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        sheets = set(wb.sheetnames)
        wb.close()
        return "List" in sheets or "Reference" in sheets
    except Exception:
        return False


def candidate_dirs(script_dir):
    """Folders to search for the master file, most-likely first."""
    home = Path.home()
    raw = [script_dir, Path.cwd(), home / "Downloads", home / "Desktop",
           home / "Documents", home / "OneDrive" / "Desktop",
           home / "OneDrive" / "Documents", home / "OneDrive"]
    out, seen = [], set()
    for d in raw:
        try:
            d = d.resolve()
        except Exception:
            continue
        if d not in seen and d.is_dir():
            seen.add(d)
            out.append(d)
    return out


def _valid(mm, dd, yyyy):
    if 1 <= mm <= 12 and 1 <= dd <= 31 and 1900 < yyyy < 2100:
        return yyyy * 10000 + mm * 100 + dd
    return None


def date_in_name(name):
    """Pull the date out of a filename like 'Barcoding_121825.xlsx' (MMDDYY)
    and return it as a comparable YYYYMMDD integer, or None if there isn't one.

    Understands MMDDYY, MMDDYYYY, and YYYYMMDD, contiguous or separated by
    -, _, /, . or spaces. If several dates appear, the newest wins.
    """
    stem = name.rsplit(".", 1)[0]
    found = []
    # separated, e.g. 12-18-25 or 12_18_2025
    for mm, dd, yy in re.findall(
            r"(?<!\d)(\d{1,2})[\-_/. ](\d{1,2})[\-_/. ](\d{2,4})(?!\d)", stem):
        yy = int(yy)
        v = _valid(int(mm), int(dd), yy if yy > 99 else 2000 + yy)
        if v:
            found.append(v)
    # contiguous 8-digit: MMDDYYYY or YYYYMMDD
    for s in re.findall(r"(?<!\d)(\d{8})(?!\d)", stem):
        found.append(_valid(int(s[:2]), int(s[2:4]), int(s[4:]))
                     or _valid(int(s[4:6]), int(s[6:]), int(s[:4])))
    # contiguous 6-digit: MMDDYY
    for s in re.findall(r"(?<!\d)(\d{6})(?!\d)", stem):
        found.append(_valid(int(s[:2]), int(s[2:4]), 2000 + int(s[4:])))
    found = [v for v in found if v]
    return max(found) if found else None


def find_master(script_dir):
    """Locate the newest local master barcoding workbook, or None.

    Chosen by the date in the filename (e.g. Barcoding_121825.xlsx -> 12/18/25);
    files without a date in the name fall back to their last-modified time.
    """
    def acceptable(f):
        n = f.name.lower()
        return not (n.startswith("~$") or "label" in n)

    def rank(f):
        d = date_in_name(f.name)
        try:
            mtime = f.stat().st_mtime
        except Exception:
            mtime = -1.0
        # dated files always beat undated ones; then newest date, then mtime
        return (1 if d is not None else 0, d or 0, mtime)

    dirs = []
    cfg = script_dir / CONFIG_NAME
    if cfg.exists():
        try:
            txt = cfg.read_text(encoding="utf-8").strip().strip('"')
        except Exception:
            txt = ""
        if txt:
            p = Path(txt)
            if p.is_file() and looks_like_master(p):
                return p
            if p.is_dir():
                dirs.append(p)
    dirs += candidate_dirs(script_dir)

    best, best_key = None, None
    # Pass 1: files named like the master; Pass 2: any xlsx that looks like it.
    for require_hint in (True, False):
        for d in dirs:
            try:
                files = list(d.glob("*.xlsx"))
            except Exception:
                continue
            for f in files:
                if not acceptable(f):
                    continue
                if require_hint and MASTER_HINT not in f.name.lower():
                    continue
                if not looks_like_master(f):
                    continue
                key = rank(f)
                if best_key is None or key > best_key:
                    best, best_key = f, key
        if best is not None:
            return best
    return None


def load_master_items(path):
    """Return {sku: description} from a master barcoding workbook.

    The hidden 'List' tab (SKU col A, description col C) is the authoritative
    superset; 'Reference' (SKU col A, desc col B) fills any gaps.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    items = {}
    if "List" in wb.sheetnames:
        for row in wb["List"].iter_rows(values_only=True):
            if row and row[0]:
                desc = row[2] if len(row) > 2 else ""
                items.setdefault(str(row[0]).strip(), str(desc or "").strip())
    if "Reference" in wb.sheetnames:
        for row in wb["Reference"].iter_rows(values_only=True):
            if row and row[0]:
                k = str(row[0]).strip()
                if k not in items:
                    desc = row[1] if len(row) > 1 else ""
                    items[k] = str(desc or "").strip()
    wb.close()
    items.pop("SKU", None)            # ignore a stray header cell if present
    return items


def sync_workbook(xlsx_path, master_items):
    """Refresh the label maker's Reference tab from master_items in place,
    preserving the Print Labels picks, formatting, and drop-down.

    Returns (added, removed, ok, reason).
    """
    import os
    wb = openpyxl.load_workbook(xlsx_path)        # keep formulas + validation
    if "Reference" not in wb.sheetnames:
        return [], [], False, "no Reference tab"
    ref = wb["Reference"]

    old = set()
    for row in ref.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and row[0]:
            old.add(str(row[0]).strip())
    new = set(master_items)
    added, removed = sorted(new - old), sorted(old - new)
    if not added and not removed:
        return [], [], True, "already current"

    if ref.max_row > 1:
        ref.delete_rows(2, ref.max_row - 1)
    for idx, (sku, desc) in enumerate(sorted(master_items.items()), start=2):
        ref.cell(idx, 1, sku)
        ref.cell(idx, 2, desc)
        ref.cell(idx, 3, f"*{sku}*")

    # Save to a temp file, then swap in -- never corrupts the original, and
    # fails cleanly if the workbook is open in Excel (file locked).
    tmp = str(xlsx_path) + ".tmp"
    try:
        wb.save(tmp)
        os.replace(tmp, xlsx_path)
    except Exception as exc:
        try:
            os.remove(tmp)
        except OSError:
            pass
        return added, removed, False, f"file in use ({exc.__class__.__name__})"
    return added, removed, True, "updated"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=WORKBOOK_NAME,
                    help="the label maker workbook to refresh")
    ap.add_argument("--master", default=None,
                    help="path to the master barcoding workbook "
                         "(auto-detected if omitted)")
    ap.add_argument("--no-pause", action="store_true",
                    help="don't wait for Enter when finished (for scripting)")
    args = ap.parse_args()

    script_dir = Path(__file__).resolve().parent
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists() and (script_dir / args.xlsx).exists():
        xlsx_path = script_dir / args.xlsx
    if not xlsx_path.exists():
        print(f"Could not find {args.xlsx} next to this script.")
        return args

    master = Path(args.master) if args.master else find_master(script_dir)
    if not master or not Path(master).is_file():
        print("Could not find a master barcoding file.")
        print("Save the latest master (e.g. in Downloads) and run again, or")
        print("pass it directly:  python update_labels.py --master <path>")
        return args

    d = date_in_name(Path(master).name)
    when = ""
    if d:
        when = f"  (dated {(d // 100) % 100:02d}/{d % 100:02d}/{d // 10000})"
    print(f"Master:  {Path(master).name}{when}")
    print(f"         {master}")
    items = load_master_items(master)
    print(f"Master has {len(items)} SKUs.")

    added, removed, ok, reason = sync_workbook(xlsx_path, items)
    if ok and (added or removed):
        print(f"\nUpdated {xlsx_path.name}: +{len(added)} new, "
              f"-{len(removed)} removed.")
        if added:
            print("  Added:   " + ", ".join(added[:12])
                  + (" ..." if len(added) > 12 else ""))
        if removed:
            print("  Removed: " + ", ".join(removed[:12])
                  + (" ..." if len(removed) > 12 else ""))
        print("\nThe drop-down is now current. Open the workbook to pick SKUs.")
    elif ok:
        print(f"\n{xlsx_path.name} is already up to date - nothing to change.")
    else:
        print(f"\nCould not update the workbook - {reason}.")
        if "file in use" in reason:
            print("Close Barcoding_Label_Maker.xlsx in Excel, then run again.")

    # Remember where the master lives for next time.
    try:
        (script_dir / CONFIG_NAME).write_text(str(Path(master).parent),
                                              encoding="utf-8")
    except Exception:
        pass
    return args


if __name__ == "__main__":
    pause = True
    try:
        parsed = main()
        pause = not parsed.no_pause
    except SystemExit:
        raise
    except Exception:
        import traceback
        print("\nSomething went wrong:\n")
        traceback.print_exc()
    if pause:
        input("\nDone. Press Enter to close...")
