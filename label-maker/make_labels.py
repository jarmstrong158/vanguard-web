"""
make_labels.py
--------------
Reads the "Print Labels" tab of the label-maker workbook and produces a
print-ready PDF of Avery labels (8 per 8.5x11 sheet, 2 columns x 4 rows).

Each label shows:
    * the SKU (bold)
    * a small, auto-truncated description (looked up from the Reference tab)
    * "BIN: <#>-<letter>" if a bin was entered
    * a scannable Code 39 barcode of the SKU (with the SKU printed under it)

Usage:
    Double-click this file, or run:
        python make_labels.py
        python make_labels.py --xlsx Barcoding_Label_Maker.xlsx --out labels.pdf

The first time it runs it auto-installs the two packages it needs
(openpyxl, reportlab) -- no separate setup step required.

No barcode font needs to be installed -- reportlab draws real Code 39 barcodes.
"""
import argparse
import importlib
import os
import subprocess
import sys
import warnings
from pathlib import Path


def ensure_packages():
    """Make sure openpyxl and reportlab are importable; install them if not.

    Lets a coworker run this with nothing but Python installed -- no need to
    distribute a .bat or run pip by hand first.
    """
    missing = []
    for pkg in ("openpyxl", "reportlab"):
        try:
            importlib.import_module(pkg)
        except ImportError:
            missing.append(pkg)
    if not missing:
        return
    print(f"First-time setup: installing {', '.join(missing)} "
          f"(needs internet, one time only)...")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install",
                               "--quiet", *missing])
    except Exception as exc:                       # pip missing / offline / etc.
        print("\nCould not install the required packages automatically.")
        print(f"  Reason: {exc}")
        print("\nFix: open a Command Prompt and run this, then try again:")
        print(f'  "{sys.executable}" -m pip install openpyxl reportlab')
        input("\nPress Enter to close...")
        sys.exit(1)
    importlib.invalidate_caches()


ensure_packages()

import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.graphics.barcode import code39
from reportlab.pdfgen import canvas

warnings.filterwarnings("ignore")

# ---- Avery sheet geometry (tweak here to match your exact label stock) ------
PAGE_W, PAGE_H = letter            # 8.5 x 11 in points (612 x 792)
COLS, ROWS = 2, 4                  # 8 labels per sheet (2 across, 4 down)

# Fixed label size: 3.5 in wide x 2 in tall
LABEL_W = 3.5 * inch
LABEL_H = 2.0 * inch
COL_GAP = 0.25 * inch              # gap between the two columns
ROW_GAP = 0.20 * inch              # gap between rows
PAD = 0.12 * inch                  # inner padding inside each label

# Center the whole grid of labels on the page
MARGIN_X = (PAGE_W - (COLS * LABEL_W + (COLS - 1) * COL_GAP)) / 2
MARGIN_TOP = (PAGE_H - (ROWS * LABEL_H + (ROWS - 1) * ROW_GAP)) / 2

# ---- Fonts ------------------------------------------------------------------
SKU_FONT = "Helvetica-Bold"
SKU_MAX_SIZE = 22          # SKU starts big...
SKU_MIN_SIZE = 7           # ...and auto-shrinks until it fits one line
PAREN_SIZE = 18           # the empty "(  )" stock-count box at the SKU's right
PAREN_GAP = 24            # blank space inside the parens for hand-writing
SKU_PAREN_GAP = 5         # tight gap between the SKU and the "(  )"
DESC_FONT, DESC_SIZE = "Helvetica", 6.5   # "very small" per request
DESC_MAX_LINES = 2
BIN_FONT, BIN_SIZE = "Helvetica-Bold", 10
BAR_HEIGHT = 0.42 * inch
BAR_WIDTH = 0.012 * inch        # preferred narrow-bar width
MIN_BAR_WIDTH = 0.0075 * inch   # scannability floor (~7.5 mil) for long SKUs
BLOCK_GAP = 5              # vertical gap between SKU / desc / bin / barcode


# ---- Auto-sync with the master barcoding workbook --------------------------
# When new codes get added to the master file (emailed around and saved on each
# PC), the label maker finds the newest local master and refreshes its own SKU
# list / descriptions on every run -- nothing to redistribute by hand.
MASTER_HINT = "barcod"            # master files have this in their filename
CONFIG_NAME = "master_location.txt"   # remembers where the master was found


def looks_like_master(path):
    """True if the workbook has the master's data tabs (List / Reference)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        sheets = set(wb.sheetnames)
        wb.close()
        return "List" in sheets or "Reference" in sheets
    except Exception:
        return False


def candidate_dirs(script_dir):
    """Folders to search for the master file, most-likely first."""
    home = Path.home()
    raw = [script_dir, Path.cwd(), home / "Downloads", home / "Desktop",
           home / "Documents", home / "OneDrive" / "Desktop",
           home / "OneDrive" / "Documents", home / "OneDrive"]
    out, seen = [], set()
    for d in raw:
        try:
            d = d.resolve()
        except Exception:
            continue
        if d not in seen and d.is_dir():
            seen.add(d)
            out.append(d)
    return out


def find_master(script_dir, workbook_path):
    """Locate the newest local master barcoding workbook, or None."""
    wb_name = Path(workbook_path).name.lower()

    def acceptable(f):
        n = f.name.lower()
        if n.startswith("~$") or n == wb_name or "label" in n:
            return False
        return True

    # An explicit path/folder in master_location.txt wins.
    dirs = []
    cfg = script_dir / CONFIG_NAME
    if cfg.exists():
        try:
            txt = cfg.read_text(encoding="utf-8").strip().strip('"')
        except Exception:
            txt = ""
        if txt:
            p = Path(txt)
            if p.is_file() and looks_like_master(p):
                return p
            if p.is_dir():
                dirs.append(p)
    dirs += candidate_dirs(script_dir)

    best, best_mtime = None, -1.0
    # Pass 1: files whose name says "barcoding" (fast, validated by structure).
    # Pass 2: any .xlsx that structurally looks like the master.
    for require_hint in (True, False):
        for d in dirs:
            try:
                files = list(d.glob("*.xlsx"))
            except Exception:
                continue
            for f in files:
                if not acceptable(f):
                    continue
                if require_hint and MASTER_HINT not in f.name.lower():
                    continue
                if not looks_like_master(f):
                    continue
                try:
                    m = f.stat().st_mtime
                except Exception:
                    continue
                if m > best_mtime:
                    best, best_mtime = f, m
        if best is not None:
            return best
    return None


def load_master_items(path):
    """Return {sku: description} from a master barcoding workbook.

    The hidden 'List' tab (SKU col A, description col C) is the authoritative
    superset; 'Reference' (SKU col A, desc col B) fills any gaps.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    items = {}
    if "List" in wb.sheetnames:
        for row in wb["List"].iter_rows(values_only=True):
            if row and row[0]:
                desc = row[2] if len(row) > 2 else ""
                items.setdefault(str(row[0]).strip(), str(desc or "").strip())
    if "Reference" in wb.sheetnames:
        for row in wb["Reference"].iter_rows(values_only=True):
            if row and row[0]:
                k = str(row[0]).strip()
                if k not in items:
                    desc = row[1] if len(row) > 1 else ""
                    items[k] = str(desc or "").strip()
    wb.close()
    items.pop("SKU", None)            # ignore a stray header cell if present
    return items


def sync_workbook(xlsx_path, master_items):
    """Refresh the label maker's Reference tab from master_items in place,
    preserving the Print Labels picks, formatting, and dropdown.

    Returns (added, removed, ok, reason).
    """
    wb = openpyxl.load_workbook(xlsx_path)        # keep formulas + validation
    if "Reference" not in wb.sheetnames:
        return [], [], False, "no Reference tab"
    ref = wb["Reference"]

    old = set()
    for row in ref.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and row[0]:
            old.add(str(row[0]).strip())
    new = set(master_items)
    added, removed = sorted(new - old), sorted(old - new)
    if not added and not removed:
        return [], [], True, "already current"

    if ref.max_row > 1:
        ref.delete_rows(2, ref.max_row - 1)
    for idx, (sku, desc) in enumerate(sorted(master_items.items()), start=2):
        ref.cell(idx, 1, sku)
        ref.cell(idx, 2, desc)
        ref.cell(idx, 3, f"*{sku}*")

    # Save to a temp file, then swap in -- never corrupts the original, and
    # fails cleanly if the workbook is open in Excel (file locked).
    tmp = str(xlsx_path) + ".tmp"
    try:
        wb.save(tmp)
        os.replace(tmp, xlsx_path)
    except Exception as exc:
        try:
            os.remove(tmp)
        except OSError:
            pass
        return added, removed, False, f"file in use ({exc.__class__.__name__})"
    return added, removed, True, "updated"


def auto_sync(script_dir, xlsx_path, master_arg):
    """Find the master and refresh the workbook. Returns the master item map
    (for live description lookup) -- empty if no master was found."""
    master = Path(master_arg) if master_arg else find_master(script_dir,
                                                             xlsx_path)
    if not master or not Path(master).is_file():
        print("(No master barcoding file found nearby - using the SKU list "
              "already in the workbook.)")
        return {}
    try:
        items = load_master_items(master)
    except Exception as exc:
        print(f"(Could not read master '{Path(master).name}': {exc})")
        return {}

    if Path(xlsx_path).exists():
        added, removed, ok, reason = sync_workbook(xlsx_path, items)
        if ok and (added or removed):
            print(f"Synced SKU list from {Path(master).name}: "
                  f"+{len(added)} new, -{len(removed)} removed.")
            if added:
                print("  New: " + ", ".join(added[:8])
                      + (" ..." if len(added) > 8 else ""))
        elif ok:
            print(f"SKU list already up to date ({Path(master).name}).")
        else:
            print(f"NOTE: could not refresh the dropdown - {reason}.")
            print("      Close the workbook in Excel, then run again, to add "
                  "the newest codes to the drop-down.")
    # Remember where the master lives for next time.
    try:
        (script_dir / CONFIG_NAME).write_text(str(Path(master).parent),
                                              encoding="utf-8")
    except Exception:
        pass
    return items


def read_rows(xlsx_path, master_items=None):
    """Return a list of dicts {sku, desc, bin} from the Print Labels tab."""
    master_items = master_items or {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Print Labels" not in wb.sheetnames:
        raise SystemExit(f"'Print Labels' tab not found in {xlsx_path}")
    ws = wb["Print Labels"]

    # SKU -> description lookup from the Reference tab (description may not be
    # cached if Excel hasn't recalculated the VLOOKUP, so look it up directly).
    desc_by_sku = {}
    if "Reference" in wb.sheetnames:
        for row in wb["Reference"].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                desc_by_sku[str(row[0]).strip()] = str(row[1] or "").strip()

    rows = []
    for r in range(5, 5 + COLS * ROWS):     # input rows B5:B12
        sku = ws.cell(r, 2).value
        if not sku:
            continue
        sku = str(sku).strip()
        bin_no = ws.cell(r, 3).value
        bin_letter = ws.cell(r, 4).value
        bin_parts = [str(p).strip() for p in (bin_no, bin_letter)
                     if p not in (None, "")]
        bin_str = "-".join(bin_parts)
        # prefer live master data, then the Reference tab, then Excel's cache
        desc = (master_items.get(sku) or desc_by_sku.get(sku)
                or str(ws.cell(r, 5).value or "").strip())
        rows.append({"sku": sku, "desc": desc, "bin": bin_str})
    return rows


def fit_lines(text, font, size, max_width, max_lines):
    """Word-wrap text to at most max_lines, truncating the last line with an
    ellipsis if it still doesn't fit ("cut short" per request)."""
    if not text:
        return []
    words = text.split()
    lines, cur = [], ""
    for w in words:
        trial = (cur + " " + w).strip()
        if stringWidth(trial, font, size) <= max_width or not cur:
            cur = trial
        else:
            lines.append(cur)
            cur = w
            if len(lines) == max_lines:
                break
    if len(lines) < max_lines and cur:
        lines.append(cur)

    # If we ran out of lines but still have text, truncate the final line.
    used = " ".join(lines)
    if used.strip() != text.strip() and lines:
        last = lines[-1]
        while last and stringWidth(last + "…", font, size) > max_width:
            last = last[:-1]
        lines[-1] = (last + "…") if last else "…"
    return lines


def fit_sku_size(sku, max_width):
    """Largest SKU font size (within bounds) that fits on one line."""
    size = SKU_MAX_SIZE
    while size > SKU_MIN_SIZE and stringWidth(sku, SKU_FONT, size) > max_width:
        size -= 0.5
    return size


def draw_label(c, x, y, item):
    """Draw one label, with every element centered horizontally and the whole
    stack centered vertically inside the label cell."""
    inner_w = LABEL_W - 2 * PAD
    cx = x + LABEL_W / 2

    # An empty "(  )" stock-count box sits tight against the SKU, and the whole
    # "SKU (  )" unit is centered. Auto-scale the SKU so the unit fits the width.
    paren_open_w = stringWidth("(", SKU_FONT, PAREN_SIZE)
    paren_box_w = paren_open_w + PAREN_GAP + stringWidth(")", SKU_FONT,
                                                         PAREN_SIZE)
    sku_region_w = inner_w - paren_box_w - SKU_PAREN_GAP

    # --- size each element up front so we can center the stack vertically ---
    sku_size = fit_sku_size(item["sku"], sku_region_w)
    sku_line_h = max(sku_size, PAREN_SIZE)
    desc_lines = fit_lines(item["desc"], DESC_FONT, DESC_SIZE, inner_w,
                           DESC_MAX_LINES)
    bin_text = f"BIN: {item['bin']}" if item["bin"] else None

    # Build the barcode to fill the label width while keeping FULL height: only
    # thin the bars as needed, never below MIN_BAR_WIDTH (scannable floor).
    probe = code39.Standard39(item["sku"], barHeight=BAR_HEIGHT,
                              barWidth=BAR_WIDTH, humanReadable=True,
                              checksum=False, quiet=False)
    fit_bw = BAR_WIDTH * inner_w / probe.width
    bw = min(BAR_WIDTH, max(MIN_BAR_WIDTH, fit_bw))
    bar = code39.Standard39(item["sku"], barHeight=BAR_HEIGHT,
                            barWidth=bw, humanReadable=True,
                            checksum=False, quiet=False)
    # last-resort uniform shrink only if an extreme SKU still overflows at floor
    bar_scale = min(1.0, inner_w / bar.width)
    bar_w = bar.width * bar_scale
    bar_bars_h = bar.barHeight * bar_scale
    bar_text_h = (bar.fontSize + 2) * bar_scale       # human-readable text below
    bar_block_h = bar_bars_h + bar_text_h

    # heights of each block (in draw order)
    blocks = [("sku", sku_line_h)]
    if desc_lines:
        blocks.append(("desc", len(desc_lines) * (DESC_SIZE + 1)))
    if bin_text:
        blocks.append(("bin", BIN_SIZE))
    blocks.append(("bar", bar_block_h))

    total_h = sum(h for _, h in blocks) + BLOCK_GAP * (len(blocks) - 1)
    cursor = y + LABEL_H / 2 + total_h / 2            # top of the stack

    for kind, h in blocks:
        if kind == "sku":
            baseline = cursor - sku_line_h
            # center the combined "SKU (  )" unit
            sku_w = stringWidth(item["sku"], SKU_FONT, sku_size)
            combo_w = sku_w + SKU_PAREN_GAP + paren_box_w
            start_x = cx - combo_w / 2
            c.setFont(SKU_FONT, sku_size)
            c.drawString(start_x, baseline, item["sku"])
            # empty "(  )" tight against the SKU for the hand-written count
            paren_x = start_x + sku_w + SKU_PAREN_GAP
            c.setFont(SKU_FONT, PAREN_SIZE)
            c.drawString(paren_x, baseline, "(")
            c.drawString(paren_x + paren_open_w + PAREN_GAP, baseline, ")")
        elif kind == "desc":
            c.setFont(DESC_FONT, DESC_SIZE)
            ty = cursor
            for line in desc_lines:
                ty -= (DESC_SIZE + 1)
                c.drawCentredString(cx, ty, line)
        elif kind == "bin":
            c.setFont(BIN_FONT, BIN_SIZE)
            c.drawCentredString(cx, cursor - BIN_SIZE, bin_text)
        elif kind == "bar":
            c.saveState()
            # bars top sits at `cursor`; text renders below the bars
            c.translate(cx - bar_w / 2, cursor - bar_bars_h)
            c.scale(bar_scale, bar_scale)
            bar.drawOn(c, 0, 0)
            c.restoreState()
        cursor -= (h + BLOCK_GAP)

    # thin guide border (comment out if your label stock is pre-cut)
    c.setLineWidth(0.25)
    c.rect(x, y, LABEL_W, LABEL_H)


def build_pdf(rows, out_path):
    c = canvas.Canvas(out_path, pagesize=letter)
    if not rows:
        c.setFont("Helvetica", 12)
        c.drawString(inch, PAGE_H - inch,
                     "No SKUs selected on the 'Print Labels' tab.")
    for i, item in enumerate(rows):
        slot = i % (COLS * ROWS)
        if i and slot == 0:
            c.showPage()
        col = slot % COLS
        row = slot // COLS
        x = MARGIN_X + col * (LABEL_W + COL_GAP)
        # rows fill top-to-bottom
        y = PAGE_H - MARGIN_TOP - (row + 1) * LABEL_H - row * ROW_GAP
        draw_label(c, x, y, item)
    c.showPage()
    c.save()


def open_file(path):
    """Open the finished PDF in the default viewer (best effort)."""
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)                       # noqa: type-checker
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception:
        pass


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="Barcoding_Label_Maker.xlsx")
    ap.add_argument("--out", default="labels_to_print.pdf")
    ap.add_argument("--master", default=None,
                    help="path to the master barcoding workbook "
                         "(auto-detected if omitted)")
    ap.add_argument("--no-sync", action="store_true",
                    help="skip refreshing the SKU list from the master")
    ap.add_argument("--no-pause", action="store_true",
                    help="don't wait for Enter when finished (for scripting)")
    ap.add_argument("--no-open", action="store_true",
                    help="don't open the PDF when finished")
    args = ap.parse_args()

    script_dir = Path(__file__).resolve().parent
    xlsx_path = Path(args.xlsx)
    # If run by double-click from elsewhere, find the workbook next to the script.
    if not xlsx_path.exists() and (script_dir / args.xlsx).exists():
        xlsx_path = script_dir / args.xlsx

    master_items = {}
    if not args.no_sync:
        master_items = auto_sync(script_dir, xlsx_path, args.master)

    rows = read_rows(str(xlsx_path), master_items)
    print(f"{len(rows)} label(s) selected:")
    for it in rows:
        print(f"  {it['sku']:<24} BIN {it['bin'] or '-':<8} {it['desc'][:40]}")
    out_path = args.out
    if not Path(out_path).is_absolute() and xlsx_path.parent != Path("."):
        out_path = str(xlsx_path.parent / args.out)
    build_pdf(rows, out_path)
    print(f"Wrote {out_path}")
    if not args.no_open:
        open_file(out_path)
    return args


if __name__ == "__main__":
    pause = True
    try:
        parsed = main()
        pause = not parsed.no_pause
    except SystemExit:
        raise
    except Exception:
        import traceback
        print("\nSomething went wrong:\n")
        traceback.print_exc()
        print("\nCommon causes: the workbook is open in Excel, the file name "
              "is different, or no SKUs were picked yet.")
    if pause:
        input("\nDone. Press Enter to close...")
