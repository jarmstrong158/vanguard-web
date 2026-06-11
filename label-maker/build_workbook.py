"""
build_workbook.py
-----------------
One-time / re-runnable builder that turns the master barcoding workbook into a
self-contained "Label Maker" workbook:

  * a "Print Labels" tab where you pick up to 8 SKUs from a drop-down
    (one page of Avery labels = 8 labels), and
  * a "Reference" tab (SKU / Description / *barcode* string) so the file is
    self-contained, plus a hidden "_SKUs" helper that powers the drop-downs.

Run this only when the master reference list changes. Day to day you just open
the generated workbook, pick SKUs, save, and run make_labels.py.

    python build_workbook.py --source "Barcoding_121825.xlsx" --out "Barcoding_Label_Maker.xlsx"
"""
import argparse
import warnings

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

warnings.filterwarnings("ignore")

LABELS_PER_PAGE = 8  # 2 columns x 4 rows on an 8.5x11 Avery sheet


def load_reference(source_path):
    """Return a list of (sku, description) from the master workbook.

    The hidden 'List' tab is the authoritative superset (SKU in col A,
    description in col C). We fall back to 'Reference' (SKU col A, desc col B)
    if 'List' is missing.
    """
    wb = openpyxl.load_workbook(source_path, data_only=True)
    items = {}
    if "List" in wb.sheetnames:
        for sku, _b, desc in wb["List"].iter_rows(values_only=True):
            if sku:
                items[str(sku).strip()] = (desc or "")
    if "Reference" in wb.sheetnames:
        for row in wb["Reference"].iter_rows(values_only=True):
            sku, desc = row[0], (row[1] if len(row) > 1 else "")
            if sku and str(sku).strip() not in items:
                items[str(sku).strip()] = (desc or "")
    # sorted, unique, no blanks
    return sorted((sku, str(desc).strip()) for sku, desc in items.items())


def build(source_path, out_path):
    items = load_reference(source_path)
    print(f"Loaded {len(items)} SKUs from {source_path}")

    wb = openpyxl.Workbook()

    # --- Print Labels (input) tab -------------------------------------------
    ws = wb.active
    ws.title = "Print Labels"
    title_font = Font(bold=True, size=14)
    hint_font = Font(italic=True, size=10, color="555555")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")

    ws["A1"] = "Print Labels"
    ws["A1"].font = title_font
    ws["A2"] = (
        "Pick up to 8 SKUs below (8 labels = one Avery sheet) and type each "
        "Bin # and Letter. Leave a row blank to skip it. Save, then run "
        "make_labels.py."
    )
    ws["A2"].font = hint_font
    ws.merge_cells("A2:E2")

    headers = {
        "A4": "#",
        "B4": "SKU (pick from list)",
        "C4": "Bin #",
        "D4": "Letter",
        "E4": "Description (auto-filled)",
    }
    for col, text in headers.items():
        ws[col] = text
        ws[col].font = hdr_font
        ws[col].fill = hdr_fill

    first_row = 5
    last_row = first_row + LABELS_PER_PAGE - 1
    for i in range(LABELS_PER_PAGE):
        r = first_row + i
        ws.cell(r, 1, i + 1).alignment = Alignment(horizontal="center")
        # Bin # and Letter are typed in by the user (E.g. 360 and D -> BIN: 360-D)
        ws.cell(r, 3).alignment = Alignment(horizontal="center")
        ws.cell(r, 4).alignment = Alignment(horizontal="center")
        # auto-fill description via lookup against the Reference tab
        ws.cell(r, 5).value = (
            f'=IFERROR(VLOOKUP(B{r},Reference!$A:$B,2,FALSE),"")'
        )
        ws.cell(r, 5).font = Font(size=9, color="555555")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 60

    # --- Reference tab (self-contained copy) --------------------------------
    ref = wb.create_sheet("Reference")
    ref["A1"], ref["B1"], ref["C1"] = "SKU", "Description", "Barcode"
    for c in ("A1", "B1", "C1"):
        ref[c].font = Font(bold=True)
    for idx, (sku, desc) in enumerate(items, start=2):
        ref.cell(idx, 1, sku)
        ref.cell(idx, 2, desc)
        ref.cell(idx, 3, f"*{sku}*")
    ref.column_dimensions["A"].width = 28
    ref.column_dimensions["B"].width = 60
    ref.column_dimensions["C"].width = 30

    # --- Hidden helper that backs the drop-down -----------------------------
    helper = wb.create_sheet("_SKUs")
    for idx, (sku, _desc) in enumerate(items, start=1):
        helper.cell(idx, 1, sku)
    helper.sheet_state = "hidden"

    n = len(items)
    name = DefinedName("SKUList", attr_text=f"_SKUs!$A$1:$A${n}")
    wb.defined_names.add(name)

    # Drop-down (data validation) on the 8 input cells
    dv = DataValidation(type="list", formula1="=SKUList", allow_blank=True,
                        showDropDown=False)
    dv.error = "Pick a SKU from the list (must exist in the Reference tab)."
    dv.errorTitle = "Unknown SKU"
    dv.prompt = "Choose a SKU from the drop-down."
    dv.promptTitle = "SKU"
    ws.add_data_validation(dv)
    dv.add(f"B{first_row}:B{last_row}")

    wb.save(out_path)
    print(f"Wrote {out_path}  (input rows B{first_row}:B{last_row})")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", default="Barcoding_121825.xlsx",
                    help="master barcoding workbook to read SKUs from")
    ap.add_argument("--out", default="Barcoding_Label_Maker.xlsx",
                    help="label-maker workbook to create")
    args = ap.parse_args()
    build(args.source, args.out)
